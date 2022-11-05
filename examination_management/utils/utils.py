import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font

import pandas as pd


def create_empty_excel(path, columns):
    empty_excel = pd.DataFrame(columns=columns)
    excel_writer = pd.ExcelWriter(path)
    empty_excel.to_excel(excel_writer, index=False)
    excel_writer.save()


def create_excel(path, data):
    excel = pd.DataFrame(data=data)
    excel_writer = pd.ExcelWriter(path)
    excel.to_excel(excel_writer, index=False)
    excel_writer.save()


def get_column(n):
    string = ["\0"] * 50
    i = 0
    while n > 0:
        rem = n % 26
        if rem == 0:
            string[i] = 'Z'
            i += 1
            n = (n // 26) - 1
        else:
            string[i] = chr((rem - 1) + ord('A'))
            i += 1
            n = n // 26
    string[i] = '\0'
    string = string[::-1]
    return "".join(string)


def get_roman(n):
    if n <= 0 or n > 10:
        return None
    r = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X']
    return r[n - 1]


def create_result_excel(path, subjects, students, semester, branch_name, batch_start, batch_end):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=(2 * len(subjects) + 9))
    cell = sheet.cell(row=1, column=1)
    cell.value = 'INDIAN INSTITUTE OF INFORMATION TECHNOLOGY, SONEPAT'
    cell.font = Font(name="Times New Roman", size=18, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=(2 * len(subjects) + 9))
    cell = sheet.cell(row=2, column=1)
    session=''
    if(semester%2==0):
        session='May./June.'
    else:
        session='Nov./Dec.'
    year=batch_start+int(0.5*semester)
    print(year)
    cell.value = f'Result Sheet for B.Tech {get_roman(semester)} Semester {branch_name} Batch {batch_start}-{batch_end} Session {session}, {year}'
    cell.font = Font(name="Times New Roman", size=16, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('A3:A6')
    cell = sheet.cell(row=3, column=1)
    cell.value = 'Sr. No.'
    cell.font = Font(name="Times New Roman", size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('B3:B6')
    cell = sheet.cell(row=3, column=2)
    cell.value = 'Roll No.'
    cell.font = Font(name="Times New Roman", size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('C3:C4')

    sheet.merge_cells('D3:D4')

    sheet.merge_cells('C5:D5')
    cell = sheet.cell(row=5, column=3)
    cell.value = 'Credits'
    cell.font = Font(name="Times New Roman", size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.cell(row=6, column=3).value = 'Name'
    sheet.cell(row=6, column=3).font = Font(name="Times New Roman", size=10)
    sheet.cell(row=6, column=3).alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
    sheet.cell(row=6, column=4).value = "Father's Name"
    sheet.cell(row=6, column=3).font = Font(name="Times New Roman", size=10)
    sheet.cell(row=6, column=4).alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

    subject_codes = list(subjects.keys())
    subject_codes.sort()
    # TO GIVE HEADINGS FOR EACH SUBJECT
    for j in range(len(subjects)):
        i = j * 2

        sheet.merge_cells(start_row=3, start_column=(i + 5), end_row=3, end_column=(i + 6))
        cell = sheet.cell(row=3, column=i + 5)
        cell.value = subjects[subject_codes[j]]['name']
        cell.font = Font(name="Times New Roman", size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        sheet.merge_cells(start_row=4, start_column=(i + 5), end_row=4, end_column=(i + 6))
        cell = sheet.cell(row=4, column=i + 5)
        cell.value = subjects[subject_codes[j]]['code']
        cell.font = Font(name="Times New Roman", size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = sheet.cell(row=5, column=i + 5)
        cell.value = subjects[subject_codes[j]]['credit']
        cell.font = Font(name="Times New Roman", size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = sheet.cell(row=6, column=i + 5)
        cell.value = 'Grade'
        cell.font = Font(name="Times New Roman", size=10)
        cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

        cell = sheet.cell(row=6, column=i + 6)
        cell.value = 'Total'
        cell.font = Font(name="Times New Roman", size=10)
        cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

    sheet.merge_cells(start_row=3, start_column=(len(subjects) * 2 + 5), end_row=6, end_column=(len(subjects) * 2 + 5))
    cell = sheet.cell(row=3, column=len(subjects) * 2 + 5)
    cell.value = 'Total Credit Sem ' + get_roman(semester)
    cell.font = Font(name="Times New Roman", size=10)
    cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

    sheet.merge_cells(start_row=3, start_column=(len(subjects) * 2 + 6), end_row=6, end_column=(len(subjects) * 2 + 6))
    cell = sheet.cell(row=3, column=len(subjects) * 2 + 6)
    cell.value = 'Total CG Sem ' + get_roman(semester)
    cell.font = Font(name="Times New Roman", size=10)
    cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

    sheet.merge_cells(start_row=3, start_column=(len(subjects) * 2 + 7), end_row=6, end_column=(len(subjects) * 2 + 7))
    cell = sheet.cell(row=3, column=len(subjects) * 2 + 7)
    cell.value = 'SGPA Sem ' + get_roman(semester)
    cell.font = Font(name="Times New Roman", size=10)
    cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

    sheet.merge_cells(start_row=3, start_column=(len(subjects) * 2 + 8), end_row=6, end_column=(len(subjects) * 2 + 8))
    cell = sheet.cell(row=3, column=len(subjects) * 2 + 8)
    cell.value = 'Serial No.'
    cell.font = Font(name="Times New Roman", size=10)
    cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

    sheet.merge_cells(start_row=3, start_column=(len(subjects) * 2 + 9), end_row=6, end_column=(len(subjects) * 2 + 9))
    cell = sheet.cell(row=3, column=len(subjects) * 2 + 9)
    cell.value = 'Re-Appears (Course Code)'
    cell.font = Font(name="Times New Roman", size=10)
    cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

    for j, (_, student) in enumerate(students.items()):
        cell = sheet.cell(row=j + 7, column=1)
        cell.value = j + 1
        cell.font = Font(name="Times New Roman", size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = sheet.cell(row=j + 7, column=2)
        cell.value = student['roll_no']
        cell.font = Font(name="Times New Roman", size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = sheet.cell(row=j + 7, column=3)
        cell.value = student['name']
        cell.font = Font(name="Times New Roman", size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')

        cell = sheet.cell(row=j + 7, column=4)
        cell.value = student['fathers_name']
        cell.font = Font(name="Times New Roman", size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')

        for i, (code, grade) in enumerate(student['grades'].items()):
            cell = sheet.cell(row=j + 7, column=5 + i * 2)
            cell.value = grade['grade']
            cell.font = Font(name="Times New Roman", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            cell = sheet.cell(row=j + 7, column=6 + i * 2)
            cell.value = grade['score']
            cell.font = Font(name="Times New Roman", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = sheet.cell(row=j + 7, column=(2 * len(subjects) + 5))
        cell.value = student['total_credit']
        cell.font = Font(name="Times New Roman", size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = sheet.cell(row=j + 7, column=(2 * len(subjects) + 6))
        cell.value = student['cg_sum']
        cell.font = Font(name="Times New Roman", size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = sheet.cell(row=j + 7, column=(2 * len(subjects) + 7))
        if len(student['reappear']):
            cell.value = 'Re-appear'
        else:
            cell.value = student['sgpa']
        cell.font = Font(name="Times New Roman", size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # TODO: Calculate Serial Number
        # cell = sheet.cell(row=j+7,column=(2*len(subjects)+8))
        # cell.value =
        # cell.font = Font(name="Times New Roman", size=11)
        # cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = sheet.cell(row=j + 7, column=(2 * len(subjects) + 9))
        cell.value = student['reappear']
        cell.font = Font(name="Times New Roman", size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(path)
